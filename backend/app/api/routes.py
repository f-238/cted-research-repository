import json
from io import BytesIO
from datetime import date
from pathlib import Path
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.exc import IntegrityError
from sqlalchemy import func, or_
from sqlalchemy.orm import Session, joinedload
from app.api.deps import require_admin, require_approved
from app.core.database import get_db
from app.core.security import create_access_token, hash_password, verify_password
from app.models.entities import AccomplishmentReport, CompletedPaper, Course, FormatCheckResult, Notification, ResearchSubmission, ReviewRemark, Template, User
from app.schemas.dto import AccomplishmentOut, CompletedPaperOut, CourseOut, DashboardStats, FacultyAccomplishmentSummaryOut, FacultyResearchItemOut, FacultyResearchResultsOut, NotificationBulkAction, NotificationBulkReadAction, NotificationCountOut, NotificationOut, ProgramCreate, ProgramOrderUpdate, ProgramYearOut, ProgramUpdate, ReportSummary, ReportTrend, ResearchBulkDelete, ReviewCreate, SearchResultOut, SearchResultsOut, SignedUrlOut, SubmissionOut, SystemSettingsPayload, TemplateOut, Token, UserCreate, UserOut
from app.services.format_checker import check_document, serialize_check
from app.services.notifications import notify, notify_admins
from app.core.config import get_settings
from app.services.storage import delete_file, signed_url, store_profile_image, store_upload
from app.services.settings import DEFAULT_SETTINGS, get_administrative_settings, save_administrative_settings

router = APIRouter(prefix="/api")


def _submission_query(db: Session):
    return db.query(ResearchSubmission).options(
        joinedload(ResearchSubmission.submitter),
        joinedload(ResearchSubmission.course),
        joinedload(ResearchSubmission.format_check),
        joinedload(ResearchSubmission.remarks),
    )


def _submission_out(item: ResearchSubmission) -> SubmissionOut:
    data = SubmissionOut.model_validate(item).model_dump()
    data["latest_remark"] = _latest_remark(item)
    return SubmissionOut(**data)


def _latest_remark(item: ResearchSubmission) -> str | None:
    latest = sorted(item.remarks, key=lambda remark: remark.created_at, reverse=True)[0] if item.remarks else None
    return latest.remarks if latest else None


def _accomplishment_query(db: Session):
    return db.query(AccomplishmentReport).options(joinedload(AccomplishmentReport.owner))


def _completed_paper_query(db: Session):
    return db.query(CompletedPaper).options(
        joinedload(CompletedPaper.program),
        joinedload(CompletedPaper.owner),
    )


def _text_mentions_user(value: str | None, user: User) -> bool:
    return bool(value and user.full_name and user.full_name.strip().lower() in value.lower())


def _ensure_report_access(item: AccomplishmentReport, user: User, write: bool = False) -> None:
    if user.role == "admin":
        return
    if not write and user.role == "faculty" and (
        _text_mentions_user(item.researcher, user) or _text_mentions_user(item.organization, user)
    ):
        return
    if item.owner_id != user.id:
        raise HTTPException(status_code=403, detail="You can only access your own records.")
    if write and item.status != "Pending":
        raise HTTPException(status_code=403, detail="Only pending records can be changed.")


def _submission_search_match(needle: str):
    return or_(
        ResearchSubmission.title.ilike(needle),
        ResearchSubmission.authors.ilike(needle),
        ResearchSubmission.adviser.ilike(needle),
        ResearchSubmission.keywords.ilike(needle),
        ResearchSubmission.abstract.ilike(needle),
        ResearchSubmission.school_year.ilike(needle),
        ResearchSubmission.course.has(or_(Course.name.ilike(needle), Course.code.ilike(needle))),
    )


def _accomplishment_search_match(needle: str):
    return or_(
        AccomplishmentReport.title.ilike(needle),
        AccomplishmentReport.researcher.ilike(needle),
        AccomplishmentReport.organization.ilike(needle),
        AccomplishmentReport.school_year.ilike(needle),
        AccomplishmentReport.category.ilike(needle),
        AccomplishmentReport.venue.ilike(needle),
    )


def _faculty_name_match(user: User):
    needle = f"%{user.full_name.strip()}%"
    return needle


def _submission_faculty_match(user: User):
    needle = _faculty_name_match(user)
    return or_(
        ResearchSubmission.authors.ilike(needle),
        ResearchSubmission.adviser.ilike(needle),
    )


def _accomplishment_faculty_match(user: User):
    needle = _faculty_name_match(user)
    return or_(
        AccomplishmentReport.researcher.ilike(needle),
        AccomplishmentReport.organization.ilike(needle),
    )


def _completed_paper_faculty_match(user: User):
    needle = _faculty_name_match(user)
    return or_(
        CompletedPaper.authors.ilike(needle),
        CompletedPaper.adviser.ilike(needle),
        CompletedPaper.remarks.ilike(needle),
    )


def _faculty_item(
    id: int,
    title: str,
    type: str,
    school_year: str,
    date_value,
    status: str,
    download_url: str | None = None,
    latest_remark: str | None = None,
    **extra,
) -> FacultyResearchItemOut:
    return FacultyResearchItemOut(
        id=id,
        title=title,
        type=type,
        school_year=school_year,
        date=str(date_value),
        status=status,
        download_url=download_url,
        latest_remark=latest_remark,
        **extra,
    )


def _configured_upload_policy(db: Session) -> tuple[set[str], int]:
    submission_settings = get_administrative_settings(db).get("submissions", {})
    file_types = submission_settings.get("accepted_file_types", {})
    allowed = set()
    if file_types.get("pdf", True):
        allowed.add(".pdf")
    if file_types.get("docx", True):
        allowed.add(".docx")
    if not allowed:
        allowed = {".pdf", ".docx"}
    return allowed, int(submission_settings.get("max_upload_size_mb") or 25)


def _publish_approved_submission(db: Session, submission: ResearchSubmission) -> None:
    if submission.submission_type == "research":
        return
    if submission.submission_type not in {"presentation", "publication", "utilization"}:
        return
    item = db.query(AccomplishmentReport).filter(AccomplishmentReport.source_submission_id == submission.id).first()
    if not item:
        item = AccomplishmentReport(source_submission_id=submission.id, report_type=submission.submission_type, owner_id=submission.submitter_id)
    item.title = submission.title
    item.researcher = submission.authors
    item.category = submission.keywords or submission.submission_type.title()
    item.organization = submission.adviser
    item.venue = submission.section or None
    item.link = None
    item.event_date = submission.created_at.date()
    item.school_year = submission.school_year
    item.status = "Approved"
    item.file_path = submission.file_path
    item.original_filename = submission.original_filename
    item.mime_type = submission.mime_type
    item.file_size = submission.file_size
    db.add(item)


def _submission_view_url(item: ResearchSubmission, user: User) -> str:
    if user.role == "admin" and item.status != "Approved":
        return "/pending-reviews"
    if item.submitter_id == user.id:
        return "/my-submissions"
    return "/repository"


def _search_result(id: int, title: str, type: str, author: str, school_year: str, status: str, view_url: str) -> SearchResultOut:
    return SearchResultOut(
        id=id,
        title=title,
        type=type,
        author=author,
        school_year=school_year,
        status=status,
        view_url=view_url,
    )


def _user_out(user: User) -> UserOut:
    avatar_url = None
    if user.profile_image_path:
        try:
            avatar_url = signed_url(
                user.profile_image_path,
                expires_in=3600,
                bucket=get_settings().supabase_profile_images_bucket,
            )
        except HTTPException:
            avatar_url = None
    data = UserOut.model_validate(user).model_dump()
    data["avatar_url"] = avatar_url
    return UserOut(**data)


@router.post("/auth/register", response_model=UserOut)
def register(payload: UserCreate, db: Session = Depends(get_db)):
    if db.query(User).filter(func.lower(User.email) == payload.email.lower()).first():
        raise HTTPException(status_code=409, detail="Email is already registered.")
    user = User(**payload.model_dump(exclude={"password"}), password_hash=hash_password(payload.password))
    db.add(user)
    db.commit()
    db.refresh(user)
    notify_admins(db, "New account registration", f"{user.full_name} registered as {user.role}.", "/users")
    db.commit()
    db.refresh(user)
    return _user_out(user)


@router.post("/auth/login", response_model=Token)
def login(email: str = Form(), password: str = Form(), db: Session = Depends(get_db)):
    user = db.query(User).filter(func.lower(User.email) == email.lower()).first()
    if not user or not verify_password(password, user.password_hash):
        raise HTTPException(status_code=401, detail="Invalid email or password.")
    if not user.is_active or user.account_status != "approved":
        raise HTTPException(status_code=403, detail="Account is not approved or active.")
    return Token(access_token=create_access_token(str(user.id)), user=_user_out(user))


@router.get("/auth/me", response_model=UserOut)
def me(user: User = Depends(require_approved)):
    return _user_out(user)


@router.get("/courses", response_model=list[CourseOut])
def courses(include_archived: bool = False, db: Session = Depends(get_db)):
    query = db.query(Course)
    if not include_archived:
        query = query.filter(Course.status != "Archived")
    return query.order_by(Course.display_order, Course.name).all()


@router.get("/settings")
def get_settings_page(db: Session = Depends(get_db), _: User = Depends(require_admin)):
    return {
        "settings": get_administrative_settings(db),
        "system_info": {
            "institution": "Jose Rizal Memorial State University",
            "campus": "Main Campus",
            "college": "College of Teacher Education",
            "system_name": "JRMSU CTED Research Repository",
            "system_version": "1.0.0",
            "database": "Supabase PostgreSQL",
            "storage": "Supabase Storage",
            "backend": "FastAPI",
            "frontend": "React + Vite",
            "hosting": "Render",
            "last_backup_date": get_administrative_settings(db).get("backup", {}).get("last_backup_date", ""),
        },
    }


@router.get("/settings/public")
def get_public_settings(db: Session = Depends(get_db)):
    settings = get_administrative_settings(db)
    return {
        "submissions": settings.get("submissions", {}),
        "academic_defaults": settings.get("academic_defaults", {}),
        "branding": settings.get("branding", {}),
    }


@router.put("/settings")
def update_settings_page(payload: SystemSettingsPayload, db: Session = Depends(get_db), _: User = Depends(require_admin)):
    return {"settings": save_administrative_settings(db, payload.settings)}


@router.post("/settings/reset")
def reset_settings_page(db: Session = Depends(get_db), _: User = Depends(require_admin)):
    return {"settings": save_administrative_settings(db, DEFAULT_SETTINGS)}


@router.get("/programs", response_model=list[CourseOut])
def list_programs(include_archived: bool = True, db: Session = Depends(get_db), _: User = Depends(require_admin)):
    query = db.query(Course)
    if not include_archived:
        query = query.filter(Course.status != "Archived")
    return query.order_by(Course.display_order, Course.name).all()


@router.post("/programs", response_model=CourseOut)
def create_program(payload: ProgramCreate, db: Session = Depends(get_db), _: User = Depends(require_admin)):
    if db.query(Course).filter(func.lower(Course.code) == payload.code.lower()).first():
        raise HTTPException(status_code=409, detail="Program code already exists.")
    item = Course(**payload.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.patch("/programs/reorder")
def reorder_programs(payload: ProgramOrderUpdate, db: Session = Depends(get_db), _: User = Depends(require_admin)):
    for program in payload.programs:
        item = db.get(Course, program.id)
        if item:
            item.display_order = program.display_order
    db.commit()
    return {"ok": True}


@router.patch("/programs/{program_id}", response_model=CourseOut)
def update_program(program_id: int, payload: ProgramUpdate, db: Session = Depends(get_db), _: User = Depends(require_admin)):
    item = db.get(Course, program_id)
    if not item:
        raise HTTPException(status_code=404, detail="Program not found.")
    existing = db.query(Course).filter(func.lower(Course.code) == payload.code.lower(), Course.id != program_id).first()
    if existing:
        raise HTTPException(status_code=409, detail="Program code already exists.")
    for key, value in payload.model_dump().items():
        setattr(item, key, value)
    db.commit()
    db.refresh(item)
    return item


@router.patch("/programs/{program_id}/archive", response_model=CourseOut)
def archive_program(program_id: int, db: Session = Depends(get_db), _: User = Depends(require_admin)):
    item = db.get(Course, program_id)
    if not item:
        raise HTTPException(status_code=404, detail="Program not found.")
    item.status = "Archived"
    db.commit()
    db.refresh(item)
    return item


@router.patch("/programs/{program_id}/restore", response_model=CourseOut)
def restore_program(program_id: int, db: Session = Depends(get_db), _: User = Depends(require_admin)):
    item = db.get(Course, program_id)
    if not item:
        raise HTTPException(status_code=404, detail="Program not found.")
    item.status = "Active"
    db.commit()
    db.refresh(item)
    return item


@router.delete("/programs/{program_id}")
def delete_program(program_id: int, db: Session = Depends(get_db), _: User = Depends(require_admin)):
    item = db.get(Course, program_id)
    if not item:
        raise HTTPException(status_code=404, detail="Program not found.")
    has_history = db.query(ResearchSubmission.id).filter(ResearchSubmission.course_id == program_id).first() or db.query(CompletedPaper.id).filter(CompletedPaper.program_id == program_id).first()
    if has_history:
        raise HTTPException(status_code=409, detail="Archive this program instead because historical records still reference it.")
    db.delete(item)
    db.commit()
    return {"ok": True}


@router.post("/programs/import")
def import_programs(file: UploadFile = File(), db: Session = Depends(get_db), _: User = Depends(require_admin)):
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Upload an Excel .xlsx file.")
    workbook = load_workbook(BytesIO(file.file.read()))
    sheet = workbook.active
    imported = 0
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        code, name, department, status, display_order = (list(row) + [None] * 5)[:5]
        if not code or not name:
            continue
        item = db.query(Course).filter(func.lower(Course.code) == str(code).lower()).first()
        if not item:
            item = Course(code=str(code).strip(), name=str(name).strip())
            db.add(item)
        item.name = str(name).strip()
        item.department = str(department).strip() if department else None
        item.status = str(status or "Active").strip()
        item.display_order = int(display_order or index)
        imported += 1
    db.commit()
    return {"ok": True, "imported": imported}


@router.get("/programs/export")
def export_programs(db: Session = Depends(get_db), _: User = Depends(require_admin)):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Programs"
    sheet.append(["Program Code", "Program Name", "Department", "Status", "Display Order"])
    for item in db.query(Course).order_by(Course.display_order, Course.name).all():
        sheet.append([item.code, item.name, item.department or "", item.status, item.display_order])
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=cted-programs.xlsx"},
    )


@router.get("/maintenance/database-backup")
def download_database_backup(db: Session = Depends(get_db), _: User = Depends(require_admin)):
    payload = {
        "settings": get_administrative_settings(db),
        "programs": [
            {"id": item.id, "code": item.code, "name": item.name, "department": item.department, "status": item.status, "display_order": item.display_order}
            for item in db.query(Course).order_by(Course.display_order, Course.name).all()
        ],
    }
    return StreamingResponse(
        BytesIO(json.dumps(payload, indent=2).encode("utf-8")),
        media_type="application/json",
        headers={"Content-Disposition": "attachment; filename=cted-database-backup.json"},
    )


@router.get("/maintenance/repository-metadata")
def download_repository_metadata(db: Session = Depends(get_db), _: User = Depends(require_admin)):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Research Metadata"
    sheet.append(["Title", "Authors", "Adviser", "Program", "School Year", "Submission Year", "Status"])
    for item in _submission_query(db).order_by(ResearchSubmission.created_at.desc()).all():
        sheet.append([item.title, item.authors, item.adviser, item.course.name if item.course else "", item.school_year, item.submission_year, item.status])
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=repository-metadata.xlsx"},
    )


@router.post("/maintenance/restore")
def restore_database_backup(file: UploadFile = File(), admin: User = Depends(require_admin)):
    return {"ok": True, "message": "Backup file received. Automated restore should be performed by the deployment administrator."}




@router.get("/dashboard/course-stats", response_model=list[DashboardStats])
def course_stats(db: Session = Depends(get_db), _: User = Depends(require_admin)):
    courses = db.query(Course).filter(Course.status != "Archived").order_by(Course.display_order, Course.name).all()
    output = []
    for course in courses:
        rows = db.query(ResearchSubmission.status, func.count(ResearchSubmission.id)).filter(
            ResearchSubmission.course_id == course.id
        ).group_by(ResearchSubmission.status).all()
        counts = dict(rows)
        output.append(DashboardStats(
            course_id=course.id,
            course_name=course.name,
            total=sum(counts.values()),
            pending=counts.get("Pending Review", 0),
            approved=counts.get("Approved", 0),
            disapproved=counts.get("Disapproved", 0),
            needs_revision=counts.get("Needs Revision", 0),
        ))
    return output


@router.get("/dashboard/report-summary", response_model=ReportSummary)
def report_summary(db: Session = Depends(get_db), user: User = Depends(require_approved)):
    if user.role not in {"admin", "faculty"}:
        raise HTTPException(status_code=403, detail="Faculty or admin access required.")
    rows = db.query(AccomplishmentReport.report_type, func.count(AccomplishmentReport.id)).group_by(AccomplishmentReport.report_type).all()
    counts = dict(rows)
    completed_count = db.query(func.count(CompletedPaper.id)).scalar() or 0
    return ReportSummary(
        presentations=counts.get("presentation", 0),
        publications=counts.get("publication", 0),
        utilizations=counts.get("utilization", 0),
        completed_papers=completed_count,
    )


@router.get("/reports/trends", response_model=list[ReportTrend])
def report_trends(db: Session = Depends(get_db), user: User = Depends(require_approved)):
    if user.role not in {"admin", "faculty"}:
        raise HTTPException(status_code=403, detail="Faculty or admin access required.")
    rows = db.query(
        AccomplishmentReport.school_year,
        AccomplishmentReport.report_type,
        func.count(AccomplishmentReport.id),
    ).group_by(AccomplishmentReport.school_year, AccomplishmentReport.report_type).all()
    grouped: dict[str, dict[str, int]] = {}
    for school_year, report_type, count in rows:
        grouped.setdefault(school_year, {})
        grouped[school_year][report_type] = count
    completed_rows = db.query(
        CompletedPaper.school_year,
        func.count(CompletedPaper.id),
    ).group_by(CompletedPaper.school_year).all()
    for school_year, count in completed_rows:
        grouped.setdefault(school_year, {})
        grouped[school_year]["completed_papers"] = count
    return [
        ReportTrend(
            school_year=school_year,
            presentations=counts.get("presentation", 0),
            publications=counts.get("publication", 0),
            utilizations=counts.get("utilization", 0),
            completed_papers=counts.get("completed_papers", 0),
        )
        for school_year, counts in sorted(grouped.items())
    ]


@router.get("/search", response_model=SearchResultsOut)
def search(q: str = "", db: Session = Depends(get_db), user: User = Depends(require_approved)):
    term = q.strip()
    empty = SearchResultsOut(research_submissions=[], presentations=[], publications=[], utilizations=[])
    if not term:
        return empty

    needle = f"%{term}%"
    submissions = _submission_query(db).filter(_submission_search_match(needle))
    if user.role != "admin":
        submissions = submissions.filter(ResearchSubmission.submitter_id == user.id)

    accomplishments = _accomplishment_query(db).filter(_accomplishment_search_match(needle))
    if user.role != "admin":
        accomplishments = accomplishments.filter(AccomplishmentReport.owner_id == user.id)

    grouped = {
        "research_submissions": [
            _search_result(
                item.id,
                item.title,
                "Research Submission",
                item.authors,
                item.school_year,
                item.status,
                _submission_view_url(item, user),
            )
            for item in submissions.order_by(ResearchSubmission.created_at.desc()).limit(30).all()
        ],
        "presentations": [],
        "publications": [],
        "utilizations": [],
    }

    for item in accomplishments.order_by(AccomplishmentReport.created_at.desc()).limit(90).all():
        key = f"{item.report_type}s"
        grouped[key].append(_search_result(
            item.id,
            item.title,
            item.report_type.title(),
            item.researcher,
            item.school_year,
            item.status,
            f"/accomplishment-reports/{item.report_type}",
        ))

    return SearchResultsOut(**grouped)


@router.get("/faculty/my-researches", response_model=FacultyResearchResultsOut)
def faculty_my_researches(db: Session = Depends(get_db), user: User = Depends(require_approved)):
    if user.role != "faculty":
        raise HTTPException(status_code=403, detail="Faculty access required.")

    submissions = _submission_query(db).filter(_submission_faculty_match(user)).order_by(ResearchSubmission.created_at.desc()).all()
    accomplishments = _accomplishment_query(db).filter(
        _accomplishment_faculty_match(user),
        AccomplishmentReport.source_submission_id.is_(None),
    ).order_by(AccomplishmentReport.event_date.desc()).all()
    completed_papers = _completed_paper_query(db).filter(
        _completed_paper_faculty_match(user),
    ).order_by(CompletedPaper.completion_date.desc()).all()

    grouped = FacultyResearchResultsOut(
        research_submissions=[],
        presentations=[],
        publications=[],
        utilizations=[],
        completed_papers=[],
    )

    submission_labels = {
        "research": "Research Submission",
        "presentation": "Presentation",
        "publication": "Publication",
        "utilization": "Utilization",
    }
    submission_groups = {
        "research": "research_submissions",
        "presentation": "presentations",
        "publication": "publications",
        "utilization": "utilizations",
    }
    for item in submissions:
        key = submission_groups.get(item.submission_type or "research", "research_submissions")
        getattr(grouped, key).append(_faculty_item(
            item.id,
            item.title,
            submission_labels.get(item.submission_type or "research", "Research Submission"),
            item.school_year,
            item.created_at.date(),
            item.status,
            f"/api/submissions/{item.id}/download",
            _latest_remark(item),
            submission_type=item.submission_type,
            authors=item.authors,
            adviser=item.adviser,
            course_id=item.course_id,
            course_name=item.course.name if item.course else None,
            section=item.section,
            submission_year=item.submission_year,
            keywords=item.keywords,
            abstract=item.abstract,
            can_edit=item.submitter_id == user.id and item.status in {"Pending Review", "Needs Revision"},
        ))

    for item in accomplishments:
        key = f"{item.report_type}s"
        getattr(grouped, key).append(_faculty_item(
            item.id,
            item.title,
            item.report_type.title(),
            item.school_year,
            item.event_date,
            item.status,
            f"/api/accomplishments/{item.id}/download" if item.file_path else None,
        ))

    for item in completed_papers:
        grouped.completed_papers.append(_faculty_item(
            item.id,
            item.title,
            "Completed Paper",
            item.school_year,
            item.completion_date,
            item.status,
            f"/api/completed-papers/{item.id}/download" if item.file_path else None,
            authors=item.authors,
            adviser=item.adviser,
            course_id=item.program_id,
            course_name=item.program.name if item.program else None,
            submission_year=item.submission_year,
            keywords=item.keywords,
            abstract=item.abstract,
        ))

    return grouped


@router.get("/faculty/accomplishment-summary", response_model=FacultyAccomplishmentSummaryOut)
def faculty_accomplishment_summary(db: Session = Depends(get_db), user: User = Depends(require_approved)):
    if user.role != "faculty":
        raise HTTPException(status_code=403, detail="Faculty access required.")

    researches = db.query(func.count(ResearchSubmission.id)).filter(
        ResearchSubmission.submission_type == "research",
        ResearchSubmission.status == "Approved",
        _submission_faculty_match(user),
    ).scalar() or 0
    rows = db.query(
        AccomplishmentReport.report_type,
        func.count(AccomplishmentReport.id),
    ).filter(
        AccomplishmentReport.status == "Approved",
        _accomplishment_faculty_match(user),
    ).group_by(AccomplishmentReport.report_type).all()
    counts = dict(rows)
    return FacultyAccomplishmentSummaryOut(
        researches=researches,
        presentations=counts.get("presentation", 0),
        publications=counts.get("publication", 0),
        utilizations=counts.get("utilization", 0),
        completed_papers=db.query(func.count(CompletedPaper.id)).filter(_completed_paper_faculty_match(user)).scalar() or 0,
    )


@router.post("/submissions", response_model=SubmissionOut)
def upload_submission(
    submission_type: str = Form("research"),
    title: str = Form(),
    authors: str = Form(),
    course_id: int = Form(),
    section: str = Form(""),
    adviser: str = Form(),
    school_year: str = Form(),
    submission_year: int = Form(),
    keywords: str = Form(),
    abstract: str = Form(),
    file: UploadFile = File(),
    db: Session = Depends(get_db),
    user: User = Depends(require_approved),
):
    if submission_type not in {"research", "presentation", "publication", "utilization"}:
        raise HTTPException(status_code=400, detail="Invalid submission type.")
    allowed_uploads, max_upload_size_mb = _configured_upload_policy(db)
    stored = store_upload(file, "research", allowed_uploads, max_upload_size_mb)
    submission = ResearchSubmission(
        submission_type=submission_type,
        title=title,
        authors=authors,
        course_id=course_id,
        section=section,
        adviser=adviser,
        school_year=school_year,
        submission_year=submission_year,
        keywords=keywords,
        abstract=abstract,
        file_path=stored.storage_key,
        original_filename=stored.original_filename,
        file_type=stored.file_type,
        mime_type=stored.mime_type,
        file_size=stored.file_size,
        submitter_id=user.id,
    )
    db.add(submission)
    db.flush()
    try:
        result = check_document(stored.temp_path)
        compliant, warnings, passed, raw = serialize_check(result)
    finally:
        Path(stored.temp_path).unlink(missing_ok=True)
    db.add(FormatCheckResult(submission_id=submission.id, is_compliant=compliant, warnings=warnings, passed_items=passed, raw_summary=raw))
    notify(db, user.id, "Submission received", f"Your {submission_type} '{title}' was submitted for review.", "/my-submissions")
    if not compliant:
        notify(db, user.id, "Format warning detected", "Review the format warnings saved with your submission.", "/my-submissions")
    notify_admins(db, "New submission uploaded", f"{user.full_name} uploaded '{title}' as {submission_type}.", "/pending-reviews")
    db.commit()
    return _submission_out(_submission_query(db).filter(ResearchSubmission.id == submission.id).first())


@router.get("/submissions", response_model=list[SubmissionOut])
def list_submissions(
    course_id: int | None = None,
    status: str | None = None,
    adviser: str | None = None,
    school_year: str | None = None,
    submission_year: int | None = None,
    mine: bool = False,
    db: Session = Depends(get_db),
    user: User = Depends(require_approved),
):
    query = _submission_query(db)
    if user.role != "admin" or mine:
        query = query.filter(ResearchSubmission.submitter_id == user.id)
    if course_id:
        query = query.filter(ResearchSubmission.course_id == course_id)
    if status:
        query = query.filter(ResearchSubmission.status == status)
    if adviser:
        query = query.filter(ResearchSubmission.adviser.ilike(f"%{adviser}%"))
    if school_year:
        query = query.filter(ResearchSubmission.school_year == school_year)
    if submission_year:
        query = query.filter(ResearchSubmission.submission_year == submission_year)
    return [_submission_out(item) for item in query.order_by(ResearchSubmission.created_at.desc()).all()]


@router.get("/repository", response_model=list[SubmissionOut])
def repository(
    search: str | None = None,
    course_id: int | None = None,
    school_year: str | None = None,
    submission_year: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(require_approved),
):
    query = _submission_query(db)
    if user.role == "admin":
        query = query.filter(
            ResearchSubmission.submission_type == "research",
            ResearchSubmission.status == "Approved",
            ResearchSubmission.visible.is_(True),
        )
    else:
        query = query.filter(ResearchSubmission.submitter_id == user.id)
    if search:
        needle = f"%{search}%"
        query = query.filter(or_(
            ResearchSubmission.title.ilike(needle),
            ResearchSubmission.authors.ilike(needle),
            ResearchSubmission.keywords.ilike(needle),
            ResearchSubmission.school_year.ilike(needle),
            ResearchSubmission.adviser.ilike(needle),
        ))
    if course_id:
        query = query.filter(ResearchSubmission.course_id == course_id)
    if school_year:
        query = query.filter(ResearchSubmission.school_year == school_year)
    if submission_year:
        query = query.filter(ResearchSubmission.submission_year == submission_year)
    return [_submission_out(item) for item in query.order_by(ResearchSubmission.title).all()]


@router.delete("/research/bulk")
def delete_research_bulk(payload: ResearchBulkDelete, db: Session = Depends(get_db), _: User = Depends(require_admin)):
    research_ids = list(dict.fromkeys(payload.research_ids))
    if not research_ids:
        return {"ok": True, "deleted": 0}

    submissions = db.query(ResearchSubmission).filter(
        ResearchSubmission.id.in_(research_ids),
        ResearchSubmission.submission_type == "research",
    ).all()
    file_paths = [submission.file_path for submission in submissions]
    for submission in submissions:
        db.query(AccomplishmentReport).filter(AccomplishmentReport.source_submission_id == submission.id).update(
            {AccomplishmentReport.source_submission_id: None},
            synchronize_session=False,
        )
        db.delete(submission)
    db.commit()
    for file_path in file_paths:
        delete_file(file_path)
    return {"ok": True, "deleted": len(submissions)}


@router.delete("/research/{research_id}")
def delete_research(research_id: int, db: Session = Depends(get_db), _: User = Depends(require_admin)):
    submission = db.get(ResearchSubmission, research_id)
    if not submission or submission.submission_type != "research":
        raise HTTPException(status_code=404, detail="Research record not found.")

    file_path = submission.file_path
    db.query(AccomplishmentReport).filter(AccomplishmentReport.source_submission_id == submission.id).update(
        {AccomplishmentReport.source_submission_id: None},
        synchronize_session=False,
    )
    db.delete(submission)
    db.commit()
    delete_file(file_path)
    return {"ok": True, "message": "Research deleted successfully."}


@router.get("/programs/{program_id}/years", response_model=list[ProgramYearOut])
def program_years(program_id: int, db: Session = Depends(get_db), _: User = Depends(require_approved)):
    rows = db.query(
        ResearchSubmission.school_year,
        ResearchSubmission.submission_year,
        ResearchSubmission.status,
        func.count(ResearchSubmission.id),
    ).filter(ResearchSubmission.course_id == program_id).group_by(
        ResearchSubmission.school_year,
        ResearchSubmission.submission_year,
        ResearchSubmission.status,
    ).all()
    grouped: dict[tuple[str, int], dict[str, int]] = {}
    for school_year, submission_year, status, count in rows:
        grouped.setdefault((school_year, submission_year), {})
        grouped[(school_year, submission_year)][status] = count
    output = []
    for (school_year, submission_year), counts in grouped.items():
        output.append(ProgramYearOut(
            course_id=program_id,
            school_year=school_year,
            submission_year=submission_year,
            total=sum(counts.values()),
            pending=counts.get("Pending Review", 0),
            approved=counts.get("Approved", 0),
            disapproved=counts.get("Disapproved", 0),
            needs_revision=counts.get("Needs Revision", 0),
        ))
    return sorted(output, key=lambda item: (item.submission_year, item.school_year), reverse=True)


@router.get("/programs/{program_id}/years/{school_year}/researches", response_model=list[SubmissionOut])
def program_year_researches(program_id: int, school_year: str, db: Session = Depends(get_db), _: User = Depends(require_approved)):
    decoded_school_year = school_year.replace("_", "-")
    query = _submission_query(db).filter(
        ResearchSubmission.course_id == program_id,
        ResearchSubmission.school_year == decoded_school_year,
    )
    return [_submission_out(item) for item in query.order_by(ResearchSubmission.created_at.desc()).all()]


@router.post("/submissions/{submission_id}/review", response_model=SubmissionOut)
def review_submission(submission_id: int, payload: ReviewCreate, db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    submission = db.get(ResearchSubmission, submission_id)
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found.")
    submission.status = payload.status
    db.add(ReviewRemark(submission_id=submission.id, reviewer_id=admin.id, status=payload.status, remarks=payload.remarks))
    if payload.status == "Approved":
        _publish_approved_submission(db, submission)
    if submission.submitter_id:
        notify(db, submission.submitter_id, f"Research {payload.status}", payload.remarks, "/my-submissions")
    db.commit()
    return _submission_out(_submission_query(db).filter(ResearchSubmission.id == submission.id).first())


@router.patch("/submissions/{submission_id}/visibility", response_model=SubmissionOut)
def update_visibility(submission_id: int, visible: bool = Form(), db: Session = Depends(get_db), _: User = Depends(require_admin)):
    submission = db.get(ResearchSubmission, submission_id)
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found.")
    submission.visible = visible
    db.commit()
    return _submission_out(_submission_query(db).filter(ResearchSubmission.id == submission.id).first())


@router.get("/submissions/{submission_id}/download", response_model=SignedUrlOut)
def download_submission(submission_id: int, db: Session = Depends(get_db), user: User = Depends(require_approved)):
    submission = db.get(ResearchSubmission, submission_id)
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found.")
    faculty_match = user.role == "faculty" and (
        _text_mentions_user(submission.authors, user) or _text_mentions_user(submission.adviser, user)
    )
    if user.role != "admin" and submission.submitter_id != user.id and not faculty_match and submission.status != "Approved":
        raise HTTPException(status_code=403, detail="You cannot access this file.")
    return SignedUrlOut(url=signed_url(submission.file_path), expires_in=300)


@router.patch("/submissions/{submission_id}", response_model=SubmissionOut)
def update_submission(
    submission_id: int,
    title: str = Form(),
    authors: str = Form(),
    course_id: int = Form(),
    section: str = Form(""),
    adviser: str = Form(),
    school_year: str = Form(),
    submission_year: int = Form(),
    keywords: str = Form(),
    abstract: str = Form(),
    file: UploadFile | None = File(None),
    db: Session = Depends(get_db),
    user: User = Depends(require_approved),
):
    submission = db.get(ResearchSubmission, submission_id)
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found.")
    previous_status = submission.status
    if user.role != "admin":
        if submission.submitter_id != user.id:
            raise HTTPException(status_code=403, detail="You can only edit your own submissions.")
        if submission.status not in {"Pending Review", "Needs Revision"}:
            raise HTTPException(status_code=403, detail="Only pending or needs revision submissions can be edited.")
    submission.title = title
    submission.authors = authors
    submission.course_id = course_id
    submission.section = section
    submission.adviser = adviser
    submission.school_year = school_year
    submission.submission_year = submission_year
    submission.keywords = keywords
    submission.abstract = abstract
    if file:
        old_key = submission.file_path
        allowed_uploads, max_upload_size_mb = _configured_upload_policy(db)
        stored = store_upload(file, "research", allowed_uploads, max_upload_size_mb)
        submission.file_path = stored.storage_key
        submission.original_filename = stored.original_filename
        submission.file_type = stored.file_type
        submission.mime_type = stored.mime_type
        submission.file_size = stored.file_size
        try:
            result = check_document(stored.temp_path)
            compliant, warnings, passed, raw = serialize_check(result)
        finally:
            Path(stored.temp_path).unlink(missing_ok=True)
        delete_file(old_key)
        if submission.format_check:
            submission.format_check.is_compliant = compliant
            submission.format_check.warnings = warnings
            submission.format_check.passed_items = passed
            submission.format_check.raw_summary = raw
        else:
            db.add(FormatCheckResult(submission_id=submission.id, is_compliant=compliant, warnings=warnings, passed_items=passed, raw_summary=raw))
    if user.role != "admin" and previous_status == "Needs Revision":
        submission.status = "Pending Review"
        notify_admins(db, "Revised submission uploaded", f"{user.full_name} resubmitted '{submission.title}' for review.", "/pending-reviews")
    db.commit()
    return _submission_out(_submission_query(db).filter(ResearchSubmission.id == submission.id).first())


@router.get("/accomplishments", response_model=list[AccomplishmentOut])
def list_accomplishments(report_type: str | None = None, mine: bool = False, db: Session = Depends(get_db), user: User = Depends(require_admin)):
    query = _accomplishment_query(db)
    if report_type:
        query = query.filter(AccomplishmentReport.report_type == report_type)
    if user.role != "admin" or mine:
        query = query.filter(AccomplishmentReport.owner_id == user.id)
    return [AccomplishmentOut.model_validate(item) for item in query.order_by(AccomplishmentReport.created_at.desc()).all()]


@router.post("/accomplishments", response_model=AccomplishmentOut)
def create_accomplishment(
    report_type: str = Form(),
    title: str = Form(),
    researcher: str = Form(),
    category: str = Form(),
    organization: str = Form(),
    event_date: date = Form(),
    school_year: str = Form(),
    status: str = Form("Pending"),
    venue: str | None = Form(None),
    link: str | None = Form(None),
    file: UploadFile | None = File(None),
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
):
    if report_type not in {"presentation", "publication", "utilization"}:
        raise HTTPException(status_code=400, detail="Invalid report type.")
    path = filename = None
    mime_type = file_size = None
    if file:
        stored = store_upload(file, "accomplishments", {".pdf", ".docx", ".doc", ".jpg", ".jpeg", ".png"})
        path = stored.storage_key
        filename = stored.original_filename
        mime_type = stored.mime_type
        file_size = stored.file_size
        Path(stored.temp_path).unlink(missing_ok=True)
    item = AccomplishmentReport(
        report_type=report_type,
        title=title,
        researcher=researcher,
        category=category,
        organization=organization,
        venue=venue,
        link=link,
        event_date=event_date,
        school_year=school_year,
        status=status,
        file_path=path,
        original_filename=filename,
        mime_type=mime_type,
        file_size=file_size,
        owner_id=user.id,
    )
    db.add(item)
    db.commit()
    return AccomplishmentOut.model_validate(_accomplishment_query(db).filter(AccomplishmentReport.id == item.id).first())


@router.patch("/accomplishments/{report_id}", response_model=AccomplishmentOut)
def update_accomplishment(
    report_id: int,
    title: str = Form(),
    researcher: str = Form(),
    category: str = Form(),
    organization: str = Form(),
    event_date: date = Form(),
    school_year: str = Form(),
    status: str = Form("Pending"),
    venue: str | None = Form(None),
    link: str | None = Form(None),
    file: UploadFile | None = File(None),
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
):
    item = db.get(AccomplishmentReport, report_id)
    if not item:
        raise HTTPException(status_code=404, detail="Record not found.")
    _ensure_report_access(item, user, write=True)
    if file:
        old_key = item.file_path
        stored = store_upload(file, "accomplishments", {".pdf", ".docx", ".doc", ".jpg", ".jpeg", ".png"})
        item.file_path = stored.storage_key
        item.original_filename = stored.original_filename
        item.mime_type = stored.mime_type
        item.file_size = stored.file_size
        Path(stored.temp_path).unlink(missing_ok=True)
        delete_file(old_key)
    item.title = title
    item.researcher = researcher
    item.category = category
    item.organization = organization
    item.venue = venue
    item.link = link
    item.event_date = event_date
    item.school_year = school_year
    item.status = status
    db.commit()
    return AccomplishmentOut.model_validate(_accomplishment_query(db).filter(AccomplishmentReport.id == item.id).first())


@router.delete("/accomplishments/{report_id}")
def delete_accomplishment(report_id: int, db: Session = Depends(get_db), user: User = Depends(require_admin)):
    item = db.get(AccomplishmentReport, report_id)
    if not item:
        raise HTTPException(status_code=404, detail="Record not found.")
    _ensure_report_access(item, user, write=True)
    delete_file(item.file_path)
    db.delete(item)
    db.commit()
    return {"ok": True}


@router.get("/accomplishments/{report_id}/download", response_model=SignedUrlOut)
def download_accomplishment(report_id: int, db: Session = Depends(get_db), user: User = Depends(require_approved)):
    item = db.get(AccomplishmentReport, report_id)
    if not item or not item.file_path:
        raise HTTPException(status_code=404, detail="File not found.")
    _ensure_report_access(item, user)
    return SignedUrlOut(url=signed_url(item.file_path), expires_in=300)


def _completed_paper_access_query(query, user: User):
    if user.role == "admin":
        return query
    if user.role == "faculty":
        return query.filter(_completed_paper_faculty_match(user))
    return query.filter(CompletedPaper.owner_id == user.id)


def _ensure_completed_paper_access(item: CompletedPaper, user: User) -> None:
    if user.role == "admin":
        return
    if user.role == "faculty" and (
        _text_mentions_user(item.authors, user)
        or _text_mentions_user(item.adviser, user)
        or _text_mentions_user(item.remarks, user)
    ):
        return
    if item.owner_id == user.id:
        return
    raise HTTPException(status_code=403, detail="You cannot access this completed paper.")


@router.get("/completed-papers", response_model=list[CompletedPaperOut])
def list_completed_papers(
    search: str | None = None,
    program_id: int | None = None,
    school_year: str | None = None,
    submission_year: int | None = None,
    adviser: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(require_approved),
):
    query = _completed_paper_access_query(_completed_paper_query(db), user)
    if search:
        needle = f"%{search}%"
        query = query.filter(or_(
            CompletedPaper.title.ilike(needle),
            CompletedPaper.authors.ilike(needle),
            CompletedPaper.adviser.ilike(needle),
            CompletedPaper.keywords.ilike(needle),
            CompletedPaper.abstract.ilike(needle),
            CompletedPaper.remarks.ilike(needle),
            CompletedPaper.school_year.ilike(needle),
            CompletedPaper.program.has(or_(Course.name.ilike(needle), Course.code.ilike(needle))),
        ))
    if program_id:
        query = query.filter(CompletedPaper.program_id == program_id)
    if school_year:
        query = query.filter(CompletedPaper.school_year == school_year)
    if submission_year:
        query = query.filter(CompletedPaper.submission_year == submission_year)
    if adviser:
        query = query.filter(CompletedPaper.adviser.ilike(f"%{adviser}%"))
    return query.order_by(CompletedPaper.completion_date.desc(), CompletedPaper.title).all()


@router.post("/completed-papers", response_model=CompletedPaperOut)
def create_completed_paper(
    title: str = Form(),
    authors: str = Form(),
    adviser: str = Form(),
    program_id: int = Form(),
    school_year: str = Form(),
    submission_year: int = Form(),
    completion_date: date = Form(),
    abstract: str = Form(),
    keywords: str = Form(),
    remarks: str | None = Form(None),
    status: str = Form("Completed"),
    owner_id: int | None = Form(None),
    file: UploadFile | None = File(None),
    uploaded_file: UploadFile | None = File(None),
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    path = filename = mime_type = None
    file_size = None
    upload = file or uploaded_file
    if upload:
        allowed_uploads, max_upload_size_mb = _configured_upload_policy(db)
        stored = store_upload(upload, "accomplishments", allowed_uploads, max_upload_size_mb)
        path = stored.storage_key
        filename = stored.original_filename
        mime_type = stored.mime_type
        file_size = stored.file_size
        Path(stored.temp_path).unlink(missing_ok=True)
    item = CompletedPaper(
        title=title,
        authors=authors,
        adviser=adviser,
        program_id=program_id,
        school_year=school_year,
        submission_year=submission_year,
        completion_date=completion_date,
        abstract=abstract,
        keywords=keywords,
        remarks=remarks,
        status=status,
        file_path=path,
        original_filename=filename,
        mime_type=mime_type,
        file_size=file_size,
        owner_id=owner_id or admin.id,
    )
    db.add(item)
    db.commit()
    return _completed_paper_query(db).filter(CompletedPaper.id == item.id).first()


@router.patch("/completed-papers/{paper_id}", response_model=CompletedPaperOut)
def update_completed_paper(
    paper_id: int,
    title: str = Form(),
    authors: str = Form(),
    adviser: str = Form(),
    program_id: int = Form(),
    school_year: str = Form(),
    submission_year: int = Form(),
    completion_date: date = Form(),
    abstract: str = Form(),
    keywords: str = Form(),
    remarks: str | None = Form(None),
    status: str = Form("Completed"),
    owner_id: int | None = Form(None),
    file: UploadFile | None = File(None),
    uploaded_file: UploadFile | None = File(None),
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    item = db.get(CompletedPaper, paper_id)
    if not item:
        raise HTTPException(status_code=404, detail="Completed paper not found.")
    upload = file or uploaded_file
    if upload:
        old_key = item.file_path
        allowed_uploads, max_upload_size_mb = _configured_upload_policy(db)
        stored = store_upload(upload, "accomplishments", allowed_uploads, max_upload_size_mb)
        item.file_path = stored.storage_key
        item.original_filename = stored.original_filename
        item.mime_type = stored.mime_type
        item.file_size = stored.file_size
        Path(stored.temp_path).unlink(missing_ok=True)
        delete_file(old_key)
    item.title = title
    item.authors = authors
    item.adviser = adviser
    item.program_id = program_id
    item.school_year = school_year
    item.submission_year = submission_year
    item.completion_date = completion_date
    item.abstract = abstract
    item.keywords = keywords
    item.remarks = remarks
    item.status = status
    item.owner_id = owner_id or item.owner_id or admin.id
    db.commit()
    return _completed_paper_query(db).filter(CompletedPaper.id == item.id).first()


@router.delete("/completed-papers/{paper_id}")
def delete_completed_paper(paper_id: int, db: Session = Depends(get_db), _: User = Depends(require_admin)):
    item = db.get(CompletedPaper, paper_id)
    if not item:
        raise HTTPException(status_code=404, detail="Completed paper not found.")
    file_path = item.file_path
    db.delete(item)
    db.commit()
    delete_file(file_path)
    return {"ok": True}


@router.get("/completed-papers/{paper_id}/download", response_model=SignedUrlOut)
def download_completed_paper(paper_id: int, db: Session = Depends(get_db), user: User = Depends(require_approved)):
    item = db.get(CompletedPaper, paper_id)
    if not item or not item.file_path:
        raise HTTPException(status_code=404, detail="File not found.")
    _ensure_completed_paper_access(item, user)
    return SignedUrlOut(url=signed_url(item.file_path), expires_in=300)


@router.get("/templates", response_model=list[TemplateOut])
def templates(db: Session = Depends(get_db)):
    return db.query(Template).order_by(Template.created_at.desc()).all()


@router.post("/templates", response_model=TemplateOut)
def create_template(title: str = Form(), instructions: str = Form(), file: UploadFile | None = File(None), db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    path = filename = None
    mime_type = file_size = None
    if file:
        allowed_uploads, max_upload_size_mb = _configured_upload_policy(db)
        stored = store_upload(file, "templates", allowed_uploads, max_upload_size_mb)
        path = stored.storage_key
        filename = stored.original_filename
        mime_type = stored.mime_type
        file_size = stored.file_size
        Path(stored.temp_path).unlink(missing_ok=True)
    template = Template(title=title, instructions=instructions, file_path=path, original_filename=filename, mime_type=mime_type, file_size=file_size, uploaded_by_id=admin.id)
    db.add(template)
    db.commit()
    db.refresh(template)
    return template


@router.get("/templates/{template_id}/download", response_model=SignedUrlOut)
def download_template(template_id: int, db: Session = Depends(get_db)):
    template = db.get(Template, template_id)
    if not template or not template.file_path:
        raise HTTPException(status_code=404, detail="Template file not found.")
    return SignedUrlOut(url=signed_url(template.file_path), expires_in=300)


@router.get("/users/me", response_model=UserOut)
def get_my_user(user: User = Depends(require_approved)):
    return _user_out(user)


@router.patch("/users/me", response_model=UserOut)
def update_my_user(
    full_name: str | None = Form(None),
    email: str | None = Form(None),
    section: str | None = Form(None),
    course_id: int | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(require_approved),
):
    if full_name is not None and full_name.strip():
        user.full_name = full_name.strip()
    if email is not None and email.strip():
        normalized_email = email.strip().lower()
        existing = db.query(User).filter(func.lower(User.email) == normalized_email, User.id != user.id).first()
        if existing:
            raise HTTPException(status_code=409, detail="Email is already registered.")
        user.email = normalized_email
    if section is not None:
        user.section = section.strip() or None
    if course_id is not None:
        user.course_id = course_id
    db.commit()
    db.refresh(user)
    return _user_out(user)


@router.post("/users/me/profile-image", response_model=UserOut)
def upload_my_profile_image(
    file: UploadFile = File(),
    db: Session = Depends(get_db),
    user: User = Depends(require_approved),
):
    old_path = user.profile_image_path
    storage_key = store_profile_image(file, user.id)
    user.profile_image_path = storage_key
    db.commit()
    db.refresh(user)
    if old_path and old_path != storage_key:
        delete_file(old_path, bucket=get_settings().supabase_profile_images_bucket)
    return _user_out(user)


@router.get("/users", response_model=list[UserOut])
def users(db: Session = Depends(get_db), _: User = Depends(require_admin)):
    return [_user_out(user) for user in db.query(User).order_by(User.created_at.desc()).all()]


@router.patch("/users/{user_id}", response_model=UserOut)
def update_user(user_id: int, account_status: str | None = Form(None), is_active: bool | None = Form(None), db: Session = Depends(get_db), _: User = Depends(require_admin)):
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found.")
    if account_status:
        user.account_status = account_status
        notify(db, user.id, "Account status updated", f"Your account is now {account_status}.")
    if is_active is not None:
        user.is_active = is_active
    db.commit()
    db.refresh(user)
    return _user_out(user)


@router.delete("/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found.")
    if user.id == admin.id:
        raise HTTPException(status_code=400, detail="You cannot delete your own account while logged in.")
    if user.role == "admin":
        admin_count = db.query(func.count(User.id)).filter(User.role == "admin").scalar() or 0
        if admin_count <= 1:
            raise HTTPException(status_code=400, detail="You cannot delete the last remaining admin account.")

    profile_image_path = user.profile_image_path
    try:
        db.query(ResearchSubmission).filter(ResearchSubmission.submitter_id == user.id).update(
            {ResearchSubmission.submitter_id: None},
            synchronize_session=False,
        )
        db.query(AccomplishmentReport).filter(AccomplishmentReport.owner_id == user.id).update(
            {AccomplishmentReport.owner_id: None},
            synchronize_session=False,
        )
        db.query(ReviewRemark).filter(ReviewRemark.reviewer_id == user.id).update(
            {ReviewRemark.reviewer_id: None},
            synchronize_session=False,
        )
        db.query(Template).filter(Template.uploaded_by_id == user.id).update(
            {Template.uploaded_by_id: None},
            synchronize_session=False,
        )
        db.query(Notification).filter(Notification.user_id == user.id).update(
            {Notification.user_id: None},
            synchronize_session=False,
        )
        db.delete(user)
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="This user account could not be deleted because one or more related records could not be archived safely.") from exc
    if profile_image_path:
        delete_file(profile_image_path, bucket=get_settings().supabase_profile_images_bucket)
    return {"ok": True}


@router.get("/notifications", response_model=list[NotificationOut])
def notifications(db: Session = Depends(get_db), user: User = Depends(require_approved)):
    return db.query(Notification).filter(Notification.user_id == user.id).order_by(Notification.created_at.desc()).all()


@router.get("/notifications/unread-count", response_model=NotificationCountOut)
def unread_notifications(db: Session = Depends(get_db), user: User = Depends(require_approved)):
    count = db.query(func.count(Notification.id)).filter(
        Notification.user_id == user.id,
        Notification.is_read.is_(False),
    ).scalar()
    return NotificationCountOut(unread_count=count or 0)


@router.delete("/notifications/bulk")
def delete_notifications_bulk(payload: NotificationBulkAction, db: Session = Depends(get_db), user: User = Depends(require_approved)):
    notification_ids = list(dict.fromkeys(payload.notification_ids))
    if not notification_ids:
        return {"ok": True, "deleted": 0}
    query = db.query(Notification).filter(
        Notification.user_id == user.id,
        Notification.id.in_(notification_ids),
    )
    deleted = query.delete(synchronize_session=False)
    db.commit()
    return {"ok": True, "deleted": deleted}


@router.patch("/notifications/bulk-read")
def update_notifications_read_state(payload: NotificationBulkReadAction, db: Session = Depends(get_db), user: User = Depends(require_approved)):
    notification_ids = list(dict.fromkeys(payload.notification_ids))
    if not notification_ids:
        return {"ok": True, "updated": 0}
    updated = db.query(Notification).filter(
        Notification.user_id == user.id,
        Notification.id.in_(notification_ids),
    ).update({Notification.is_read: payload.is_read}, synchronize_session=False)
    db.commit()
    return {"ok": True, "updated": updated}


@router.delete("/notifications/{notification_id}")
def delete_notification(notification_id: int, db: Session = Depends(get_db), user: User = Depends(require_approved)):
    item = db.get(Notification, notification_id)
    if not item or item.user_id != user.id:
        raise HTTPException(status_code=404, detail="Notification not found.")
    db.delete(item)
    db.commit()
    return {"ok": True}


@router.patch("/notifications/{notification_id}/read", response_model=NotificationOut)
def read_notification(notification_id: int, db: Session = Depends(get_db), user: User = Depends(require_approved)):
    item = db.get(Notification, notification_id)
    if not item or item.user_id != user.id:
        raise HTTPException(status_code=404, detail="Notification not found.")
    item.is_read = True
    db.commit()
    return item
